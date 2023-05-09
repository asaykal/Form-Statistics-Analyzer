from flask import Flask, render_template, request, jsonify, redirect, url_for, make_response
from openpyxl import Workbook, load_workbook
from datetime import date
import os, statistics, base64, openai, yaml, matplotlib, io, csv

matplotlib.use('Agg')  

import matplotlib.pyplot as plt
from io import BytesIO
from collections import Counter

app = Flask(__name__)

class formApp:
    def __init__(self):
        self.time = ""
        self.mission = ""
        self.goal = ""
        self.goal_point = 0
        self.pleasure_point = 0
        self.notes = ""

def save_to_excel(time, mission, goal, goal_point, pleasure_point, notes):
    today = date.today()
    file_name = f"form_results_{today}.xlsx"

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Mission", "Goal", "Goal Point", "Pleasure Point", "Notes"])

    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1).value = time
    ws.cell(row=next_row, column=2).value = mission
    ws.cell(row=next_row, column=3).value = goal
    ws.cell(row=next_row, column=4).value = goal_point
    ws.cell(row=next_row, column=5).value = pleasure_point
    ws.cell(row=next_row, column=6).value = notes

    wb.save(file_name)

@app.route('/')
def index():
    file_name = f"form_results_{date.today()}.xlsx"

    if not os.path.exists(file_name):
        return render_template('index.html', error='No data available')

    wb = load_workbook(file_name)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2):
        time = row[0].value
        mission = row[1].value
        goal = row[2].value
        goal_point = row[3].value
        pleasure_point = row[4].value
        notes = row[5].value

        data.append({
            'time': time,
            'mission': mission,
            'goal': goal,
            'goal_point': goal_point,
            'pleasure_point': pleasure_point,
            'notes': notes
        })

    return render_template('index.html', data=data)

@app.route('/delete', methods=['POST'])
def delete_entry():
    time = request.form.get('time')
    mission = request.form.get('mission')
    goal = request.form.get('goal')
    goal_point = int(request.form.get('goal_point'))
    pleasure_point = int(request.form.get('pleasure_point'))
    notes = request.form.get('notes')

    file_name = f"form_results_{date.today()}.xlsx"

    if not os.path.exists(file_name):
        return redirect(url_for('index'))

    wb = load_workbook(file_name)
    ws = wb.active

    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        row_time = row[0].value
        row_mission = row[1].value
        row_goal = row[2].value
        row_goal_point = row[3].value
        row_pleasure_point = row[4].value
        row_notes = row[5].value

        if row_time == time and row_mission == mission and row_goal == goal and row_goal_point == goal_point and row_pleasure_point == pleasure_point and row_notes == notes:
            rows_to_delete.append(row)

    for row in rows_to_delete:
        ws.delete_rows(row[0].row)

    wb.save(file_name)

    return redirect(url_for('index'))

@app.route('/form', methods=['POST'])
def form_endpoint():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid request data'})

    time = data.get('time')
    mission = data.get('mission')
    goal = data.get('goal')
    goal_point = data.get('goal_point')
    pleasure_point = data.get('pleasure_point')
    notes = data.get('notes')

    if not time or not mission or not goal:
        return jsonify({'error': 'Missing required form data'})

    try:
        goal_point = int(goal_point)
        pleasure_point = int(pleasure_point)
    except ValueError:
        return jsonify({'error': 'Invalid form data type'})

    form_app = formApp() 

    form_app.time = time
    form_app.mission = mission
    form_app.goal = goal
    form_app.goal_point = goal_point
    form_app.pleasure_point = pleasure_point
    form_app.notes = notes

    save_to_excel(form_app.time, form_app.mission, form_app.goal, form_app.goal_point, form_app.pleasure_point, form_app.notes)

    return jsonify({'response': 'Form is submitted'})

with open("config.yaml", "r") as f:
    config = yaml.safe_load(f)

api_key = config["openai"]["api_key"]
openai.api_key = api_key

def analyze_statistics(times, mission_names, goal_points, pleasure_points, notes):
    times_str = '\n'.join(times)
    mission_names_str = '\n'.join(mission_names)
    goal_points_str = '\n'.join(str(point) for point in goal_points)
    pleasure_points_str = '\n'.join(str(point) for point in pleasure_points)
    notes_str = '\n'.join(notes)

    prompt = f"""
    
    Today is {date.today()}.
    
    Times:
    {times_str}
    
    (Times mean : If 15 - 16 = mission has done in 15:00 - 16:00 in 24 hour format )
    
    Mission Names:
    {mission_names_str}

    Goal Points:
    {goal_points_str}

    Pleasure Points:
    {pleasure_points_str}

    Notes:
    {notes_str}

    Analyze the statistics and provide insights.
    And give activity, topic or things(book, movie, food etc.) sugesstions based on the statistics.
    1-

    2-

    3-

    4-
    
    5-
    """

    completion = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages = [
                    {"role" : "system", "content" :   
                                                    """ 
                                                    You are an Statistician and Psychologist. You analyze the daily statistics of people and provide insights. You also give them advice on how to improve their lives. And similar missions that person had good performance.  
                                                    """ 
                                                    },
                    {"role" : "user", "content" : f" {prompt} "}, 
                    {"role" : "assistant", "content" : "Assistant : "}
                            ],
                n=1,
                stop=None, 
                temperature=0.4,
            )
    analyzed_statistics = completion['choices'][0]['message']['content']

    return analyzed_statistics

def analyzed_statistics_method():
    file_name = f"form_results_{date.today()}.xlsx"
    wb = load_workbook(file_name)
    ws = wb.active

    mission_names = [cell.value for cell in ws['B'][1:]]
    times = [cell.value for cell in ws['A'][1:]]
    goal_points = [cell.value for cell in ws['D'][1:]]
    pleasure_points = [cell.value for cell in ws['E'][1:]]
    notes = [cell.value for cell in ws['F'][1:]]

    analyzed_statistics = analyze_statistics(times, mission_names, goal_points, pleasure_points, notes)
    return analyzed_statistics

message_history = []
analyzed_statistics = analyzed_statistics_method() 
#message_history.append({'role': 'system', 'content': f"Statistics Analysis:\n{analyzed_statistics}\nThese data belongs to a 21 year old person who is in deep depression condition for about 3 years. And this data form created for suggession on psychologist suggestion."})

@app.route('/talk_about_statistics', methods=['GET', 'POST'])
def talk_about_statistics_endpoint():
    file_name = f"form_results_{date.today()}.xlsx"
    wb = load_workbook(file_name)
    ws = wb.active

    mission_names = [cell.value for cell in ws['B'][1:]]
    times = [cell.value for cell in ws['A'][1:]]
    goal_points = [cell.value for cell in ws['D'][1:]]
    pleasure_points = [cell.value for cell in ws['E'][1:]]
    notes = [cell.value for cell in ws['F'][1:]]
    
    times_str = '\n'.join(times)
    mission_names_str = '\n'.join(mission_names)
    goal_points_str = '\n'.join(str(point) for point in goal_points)
    pleasure_points_str = '\n'.join(str(point) for point in pleasure_points)
    notes_str = '\n'.join(notes)

    message_history.append({
        'role': 'user', 'content': f"""
    
        Today is {date.today()}.
        
        Times:
        {times_str}
        
        Mission Names:
        {mission_names_str}

        Goal Points:
        {goal_points_str}

        Pleasure Points:
        {pleasure_points_str}

        Notes:
        {notes_str}
        
        Statistics Analysis:\n{analyzed_statistics}\n
        
        These data belongs to a 21 year old person who is in deep depression condition for about 3 years. 
        And this data form created on my psychologist terapists's suggestion.
        """
        })
    
    if request.method == 'POST':
        user_input = request.form.get('user_input')
        if user_input:
            message_history.append({'role': 'user', 'content': user_input})
            completion = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {'role': 'system', 'content': 'You are a Statistician and Psychologist...'},
                    *message_history,
                    {'role': 'assistant', 'content': 'Assistant: '}
                ],
                n=1,
                stop=None,
                temperature=0.4,
            )
            response = completion['choices'][0]['message']['content']
            message_history.append({'role': 'assistant', 'content': response})

    return render_template('chat.html', message_history=message_history)

@app.route('/send_message', methods=['POST'])
def send_message():
    message = request.form.get('message')
    if message:
        message_history.append({'role': 'user', 'content': message})
        completion = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {'role': 'system', 'content': 'You are a Statistician and Psychologist...'},
                *message_history,
                {'role': 'assistant', 'content': 'Assistant: '}
            ],
            n=1,
            stop=None,
            temperature=0.4,
        )
        response = completion['choices'][0]['message']['content']
        message_history.append({'role': 'assistant', 'content': response})

    return jsonify({'response': response})

@app.route('/export_history')
def export_history():
    output = io.StringIO()  
    writer = csv.writer(output)
    for message in message_history:
        writer.writerow([message['role']])
        writer.writerow([message['content']])
        writer.writerow([''])
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=history.csv'
    response.headers['Content-Type'] = 'text/csv'
    return response

@app.route('/statistics')
def statistics_endpoint():
    file_name = f"form_results_{date.today()}.xlsx"

    if not os.path.exists(file_name):
        return render_template('statistics.html', error='No data available')

    wb = load_workbook(file_name)
    ws = wb.active
    
    mission_names = [cell.value for cell in ws['B'][1:]]
    times = [cell.value for cell in ws['A'][1:]]

    goal_points = [cell.value for cell in ws['D'][1:]]
    pleasure_points = [cell.value for cell in ws['E'][1:]]
    
    goal_points = [int(point) for point in goal_points if point is not None]
    pleasure_points = [int(point) for point in pleasure_points if point is not None]

    mission_counts = Counter(mission_names)
    unique_missions = list(mission_counts.keys())
    mission_occurrences = list(mission_counts.values())

    if not goal_points or not pleasure_points:
        return render_template('statistics.html', error='No data available')

    goal_point_avg = statistics.mean(goal_points)
    pleasure_point_avg = statistics.mean(pleasure_points)
    goal_point_max = max(goal_points)
    pleasure_point_max = max(pleasure_points)

    plt.figure(figsize=(8, 6))
    plt.bar(range(len(goal_points)), goal_points)
    plt.xticks(range(len(goal_points)), [f'{mission}\n{time}' for mission, time in zip(mission_names, times)], rotation='vertical')
    plt.xlabel('Yapılan ve Zaman Aralığı')
    plt.ylabel('Hedefe Ulaşma Puanı')
    plt.title('Hedef Puanı Dağılımı')
    plt.tight_layout()

    goal_points_buffer = BytesIO()
    plt.savefig(goal_points_buffer, format='png')
    goal_points_buffer.seek(0)
    goal_points_data = base64.b64encode(goal_points_buffer.getvalue()).decode('utf-8')

    plt.figure(figsize=(8, 6))
    plt.bar(range(len(pleasure_points)), pleasure_points)
    plt.xticks(range(len(pleasure_points)), [f'{mission}\n{time}' for mission, time in zip(mission_names, times)], rotation='vertical')
    plt.xlabel('Yapılan ve Zaman Aralığı')
    plt.ylabel('Keyif Alma Puanı')
    plt.title('Keyif Puanı Dağılımı')
    plt.tight_layout()

    pleasure_points_buffer = BytesIO()
    plt.savefig(pleasure_points_buffer, format='png')
    pleasure_points_buffer.seek(0)
    pleasure_points_data = base64.b64encode(pleasure_points_buffer.getvalue()).decode('utf-8')

    plt.figure(figsize=(8, 6))
    plt.pie(mission_occurrences, labels=unique_missions, autopct='%1.1f%%')
    plt.title('Mission Occurrences')
    plt.tight_layout()

    mission_occurrences_buffer = BytesIO()
    plt.savefig(mission_occurrences_buffer, format='png')
    mission_occurrences_buffer.seek(0)
    mission_occurrences_data = base64.b64encode(mission_occurrences_buffer.getvalue()).decode('utf-8')
    
    #notes = [cell.value for cell in ws['F'][1:]]

    return render_template('statistics.html', goal_point_avg=goal_point_avg,
                        pleasure_point_avg=pleasure_point_avg,
                        goal_point_max=goal_point_max,
                        pleasure_point_max=pleasure_point_max,
                        goal_points_data=goal_points_data,
                        pleasure_points_data=pleasure_points_data,
                        mission_occurrences_data=mission_occurrences_data,
                        )                        #analyzed_statistics=analyzed_statistics

@app.route('/clear_history', methods=['POST'])
def clear_history():
    global message_history
    message_history = []
    return redirect('/talk_about_statistics')

if __name__ == '__main__':
    app.run(port=8080)
